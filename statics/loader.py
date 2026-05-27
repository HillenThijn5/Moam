# static_info/input.py
#haalt esg scores en benchmarks per onderliggende op


import sys
from pathlib import Path
from openpyxl import load_workbook


def _exe_root() -> Path:
    """Projectroot: map van de exe als die draait, anders de bronroot van MoamProject."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parents[1]


_STATIC_PATH = _exe_root() / "statics" / "static_sheet.xlsx"

# Caches op moduleniveau — gevuld bij de eerste aanroep en daarna hergebruikt
_benchmark_cache: dict | None = None
_esg_cache: dict | None = None
_adviser_cache: dict | None = None


def load_benchmark_map() -> dict[str, dict[str, str]]:
    """
    Leest tabblad 'Benchmarks' uit STATIC_DATA_PATH.

    Verwachte headers (rij 1):
      ticker | primary_benchmark | fallback_benchmark
    """
    global _benchmark_cache
    if _benchmark_cache is not None:
        return _benchmark_cache
    wb = load_workbook(str(_STATIC_PATH), data_only=True)
    ws = wb["Benchmarks"]

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    # Zoek vereiste kolommen op headernaam
    def col_idx(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            return -1

    i_ticker = col_idx("BENCHMARK ID")
    i_primary = col_idx("BENCHMARK NAME")
    i_fallback = col_idx("ALTERNATIVE BENCHMARK NAME")

    out: dict[str, dict[str, str]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue

        ticker = (row[i_ticker] if i_ticker >= 0 else None) or ""
        ticker = str(ticker).strip()
        if not ticker:
            continue

        primary = (row[i_primary] if i_primary >= 0 else "") if i_primary >= 0 else ""
        primary = str(primary).strip() if primary is not None else ""
        if not primary:
            # Geen benchmarknaam → overslaan (bijv. EEM UP is een fonds zonder benchmark)
            continue

        fallback = (row[i_fallback] if i_fallback >= 0 else "") if i_fallback >= 0 else ""

        out[ticker] = {
            "primary": primary,
            "fallback": str(fallback).strip() if fallback is not None else "",
        }

    wb.close()
    _benchmark_cache = out
    return out

def load_esg_map() -> dict[str, int]:
    global _esg_cache
    if _esg_cache is not None:
        return _esg_cache
    wb = load_workbook(str(_STATIC_PATH), data_only=True)
    ws = wb["ESG"]

    headers = [str(c.value).strip().upper() if c.value else "" for c in ws[1]]

    def col(name: str):
        try:
            return headers.index(name)
        except ValueError:
            return None

    i_ticker = col("TICKER")
    i_score = col("SCORE")

    out = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue

        ticker = str(row[i_ticker]).strip() if i_ticker is not None else ""
        if not ticker:
            continue

        try:
            score = int(row[i_score]) if i_score is not None else None
        except (TypeError, ValueError):
            score = None

        if score is not None:
            out[ticker] = score

    wb.close()
    _esg_cache = out
    return out


def load_adviser_map() -> dict[str, str]:
    """
    Leest tabblad 'Advisers' uit static_sheet.xlsx.

    Verwachte kolommen (geen headerrij vereist, lege regels worden overgeslagen):
      A: adviseurnaam   B: helpernaam (CC)

    Bouwt ook een voornaamindex, zodat gedeeltelijke namen (bijv. "Bauke")
    kunnen worden opgelost naar volledige namen (bijv. "Bauke van Sluis") voor CC-opzoeking.
    """
    global _adviser_cache, _adviser_first_name_map
    if _adviser_cache is not None:
        return _adviser_cache
    wb = load_workbook(str(_STATIC_PATH), data_only=True)
    ws = wb["advisers"]

    out: dict[str, str] = {}
    first_map: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        adviser = str(row[0]).strip()
        helper  = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if adviser:
            out[adviser] = helper
            first = adviser.split()[0].lower()
            # Alleen koppelen als het eenduidig is (eerste voorkomen wint)
            if first not in first_map:
                first_map[first] = adviser

    wb.close()
    _adviser_cache = out
    _adviser_first_name_map = first_map
    return out


_adviser_first_name_map: dict[str, str] = {}


def resolve_adviser_name(name: str) -> str:
    """Los een gedeeltelijke adviseurnaam (bijv. 'Bauke') op naar de volledige naam.

    Geeft de oorspronkelijke naam terug als er geen match is.
    """
    name = name.strip()
    if not name:
        return name
    # Al een volledige match
    adviser_map = load_adviser_map()
    if name in adviser_map:
        return name
    # Probeer opzoeken op voornaam
    full = _adviser_first_name_map.get(name.split()[0].lower())
    return full if full else name


# ── hulpfuncties ─────────────────────────────────────────────────────────────

def _col(headers: list[str], name: str) -> int:
    try:
        return headers.index(name.upper())
    except ValueError:
        return -1


def _str(val) -> str:
    return str(val).strip() if val is not None else ""


def _or_none(val) -> str | None:
    s = _str(val)
    return s if s else None


# ── Onderliggende waarden ─────────────────────────────────────────────────────

_underlyings_cache: dict | None = None


def load_underlyings() -> dict:
    """
    Leest tabblad 'Benchmarks' voor metadata van onderliggende waarden.
    Gebruikt kolommen: BENCHMARK ID (ticker) | FULL_NAME | ALIAS

    Alle rijen met een BENCHMARK ID worden als geldige onderliggende waarden behandeld,
    ongeacht of ze een benchmark hebben (bijv. EEM UP is een fonds
    zonder BENCHMARK NAME, maar wel een geldige onderliggende waarde).

    Geeft terug {
        'list':       [ticker, ...],          # geordend; stuurt GUI-dropdowns aan
        'full_names': {ticker: full_name},    # weergave in marketingmail-Excel
        'aliases':    {ticker: alias},        # korte namen voor onderwerp / titel
    }
    """
    global _underlyings_cache
    if _underlyings_cache is not None:
        return _underlyings_cache

    wb = load_workbook(str(_STATIC_PATH), data_only=True)
    try:
        ws = wb["Benchmarks"]
    except KeyError:
        wb.close()
        _underlyings_cache = {"list": [], "full_names": {}, "aliases": {}}
        return _underlyings_cache

    headers  = [_str(c.value).upper() for c in ws[1]]
    i_tick   = _col(headers, "BENCHMARK ID")
    i_bname  = _col(headers, "BENCHMARK NAME")
    i_full   = _col(headers, "FULL_NAME")
    i_alias  = _col(headers, "ALIAS")

    tickers, full_names, aliases = [], {}, {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        ticker = _str(row[i_tick]) if i_tick >= 0 else ""
        if not ticker:
            continue
        tickers.append(ticker)
        # FULL_NAME krijgt voorrang; val voor gewone indices terug op BENCHMARK NAME
        explicit = _str(row[i_full]) if i_full >= 0 else ""
        fallback = _str(row[i_bname]) if i_bname >= 0 else ""
        display  = explicit or fallback
        if display:
            full_names[ticker] = display
        if i_alias >= 0 and row[i_alias]:
            aliases[ticker] = _str(row[i_alias])

    wb.close()
    _underlyings_cache = {"list": tickers, "full_names": full_names, "aliases": aliases}
    return _underlyings_cache


# ── Product-URL's ─────────────────────────────────────────────────────────────

_product_urls_cache: dict | None = None


def load_product_urls() -> dict:
    """
    Leest tabblad 'ProductURLs'.
    Headers (rij 1):
      PRODUCT_TYPE | BROCHURE_LABEL | BROCHURE_URL | VIDEO_URL
                   | VIDEO_LABEL_MARKETING | VIDEO_LABEL_PC

    Geeft terug {
        product_type: {
            'brochure_label':       str,
            'brochure_url':         str,
            'video_url':            str | None,
            'video_label':          str | None,   # marketing (Nederlands)
            'video_label_pc':       str | None,   # PC-mail (Engels)
        }, ...
    }
    """
    global _product_urls_cache
    if _product_urls_cache is not None:
        return _product_urls_cache

    wb = load_workbook(str(_STATIC_PATH), data_only=True)
    try:
        ws = wb["ProductURLs"]
    except KeyError:
        wb.close()
        _product_urls_cache = {}
        return _product_urls_cache

    headers = [_str(c.value).upper() for c in ws[1]]
    i_type  = _col(headers, "PRODUCT_TYPE")
    i_bl    = _col(headers, "BROCHURE_LABEL")
    i_bu    = _col(headers, "BROCHURE_URL")
    i_vu    = _col(headers, "VIDEO_URL")
    i_vlm   = _col(headers, "VIDEO_LABEL_MARKETING")
    i_vlp   = _col(headers, "VIDEO_LABEL_PC")

    out: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        pt = _str(row[i_type]) if i_type >= 0 else ""
        if not pt:
            continue
        out[pt] = {
            "brochure_label": _str(row[i_bl])  if i_bl  >= 0 else "",
            "brochure_url":   _str(row[i_bu])  if i_bu  >= 0 else "",
            "video_url":      _or_none(row[i_vu]  if i_vu  >= 0 else None),
            "video_label":    _or_none(row[i_vlm] if i_vlm >= 0 else None),
            "video_label_pc": _or_none(row[i_vlp] if i_vlp >= 0 else None),
        }

    wb.close()
    _product_urls_cache = out
    return out


# ── PARP-datums ───────────────────────────────────────────────────────────────

_parp_cache: dict | None = None


def load_parp_dates() -> dict[str, str]:
    """
    Leest tabblad 'PARP'. Headers (rij 1): PRODUCT_TYPE | PARP_DATE
    Bewaart datums als platte tekst, bijv. '24 September 2025'.
    Geeft terug {product_type: date_string}.
    """
    global _parp_cache
    if _parp_cache is not None:
        return _parp_cache

    wb = load_workbook(str(_STATIC_PATH), data_only=True)
    try:
        ws = wb["PARP"]
    except KeyError:
        wb.close()
        _parp_cache = {}
        return _parp_cache

    headers = [_str(c.value).upper() for c in ws[1]]
    i_type = _col(headers, "PRODUCT_TYPE")
    i_date = _col(headers, "PARP_DATE")

    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        pt = _str(row[i_type]) if i_type >= 0 else ""
        if not pt:
            continue
        val = row[i_date] if i_date >= 0 else None
        if val is not None:
            from datetime import date as _date, datetime as _dt
            if isinstance(val, (_date, _dt)):
                # Formatteer als "24 September 2025" (zonder voorloopnul)
                out[pt] = val.strftime("%d %B %Y").lstrip("0")
            else:
                out[pt] = _str(val)

    wb.close()
    _parp_cache = out
    return out


# ── Prospectus-URL's ──────────────────────────────────────────────────────────

_prospectus_cache: dict | None = None


def load_prospectus_urls() -> dict[str, tuple[str, str]]:
    """
    Leest tabblad 'ProspectusURLs'. Headers (rij 1): CODE | LABEL | URL
    Geeft terug {code: (label, url)}.
    """
    global _prospectus_cache
    if _prospectus_cache is not None:
        return _prospectus_cache

    wb = load_workbook(str(_STATIC_PATH), data_only=True)
    try:
        ws = wb["ProspectusURLs"]
    except KeyError:
        wb.close()
        _prospectus_cache = {}
        return _prospectus_cache

    headers = [_str(c.value).upper() for c in ws[1]]
    i_code  = _col(headers, "CODE")
    i_label = _col(headers, "LABEL")
    i_url   = _col(headers, "URL")

    out: dict[str, tuple[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        code = _str(row[i_code]) if i_code >= 0 else ""
        if not code:
            continue
        label = _str(row[i_label]) if i_label >= 0 else code
        url   = _str(row[i_url])   if i_url   >= 0 else ""
        out[code] = (label or code, url)

    wb.close()
    _prospectus_cache = out
    return out
