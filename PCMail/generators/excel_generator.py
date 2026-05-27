# generators/excel_generator.py

from pathlib import Path
from openpyxl import load_workbook
from PCMail.config.paths import EXCEL_TM_TEMPLATE_PATH, TEMP_DIR

from openpyxl.styles import PatternFill

def render_target_market(b24_value: str, series: str, sheet_name: str | None = None) -> str:
    wb = load_workbook(str(EXCEL_TM_TEMPLATE_PATH))

    # Kies het tabblad
    ws = wb[sheet_name] if sheet_name else wb.active

    # ✅ Werk die ene cel bij
    cell = ws["B24"]
    cell.value = b24_value
    apply_target_market_fill(cell, b24_value)

    # ✅ Verwijder alle andere tabbladen (behoudt de opmaak van het overgebleven tabblad)
    for s in list(wb.worksheets):
        if s.title != ws.title:
            wb.remove(s)

    # Optioneel: geef het overgebleven tabblad een nette naam
    ws.title = "Target Market"

    out_path = Path(TEMP_DIR) / f"Target Market - series {series}.xlsx"
    wb.save(str(out_path))
    return str(out_path)

def apply_target_market_fill(cell, value: str):
    if not value:
        return

    v = value.strip().upper()

    if "YES" in v:
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="FF00B050",
            end_color="FF00B050",
        )

    elif "NEUTRAL" in v:
        cell.fill = PatternFill(
            fill_type="solid",
            start_color="FFFFC000",
            end_color="FFFFC000",
        )

