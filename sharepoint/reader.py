# sharepoint/reader.py
"""
Leest (en ververst optioneel) het SharePoint-Excelbestand "New Notes Summary".
Het bestand wordt lokaal gecachet via OneDrive-sync, dus we kopiëren het naar een tijdelijk pad
om permissiefouten te voorkomen wanneer het bestand ook in Excel is geopend.

Verversstrategie: roep start_background_refresh() één keer aan bij het opstarten van de app.
Het selectiedialoogvenster leest alleen; het start nooit nog een refresh.
"""
from __future__ import annotations

import ctypes
import msvcrt
import os
import shutil
import tempfile
import threading
import time
from pathlib import Path

import openpyxl

SHEET_NAME = "New Notes Summary"

# Status op moduleniveau zodat de app de opstartrefresh kan controleren / afwachten
_refresh_thread: threading.Thread | None = None
_refresh_done   = threading.Event()
_refresh_error: str = ""


# ── Bestandskopie met gedeelde leestoegang (omzeilt OneDrive-/Excel-locks) ──

_FILE_SHARE_READ  = 0x00000001
_FILE_SHARE_WRITE = 0x00000002
_OPEN_EXISTING    = 3
_GENERIC_READ     = 0x80000000


def _shared_copy(src: Path, dst: Path) -> None:
    """
    Kopieer *src* → *dst* met gedeelde leestoegang van Windows.
    Dit werkt zelfs wanneer OneDrive of Excel een lock op het bronbestand heeft.
    Valt terug op shutil.copy2 op niet-Windows-systemen of als de API-aanroep onverwacht mislukt.
    """
    try:
        kernel32 = ctypes.windll.kernel32
        handle = kernel32.CreateFileW(
            str(src),
            _GENERIC_READ,
            _FILE_SHARE_READ | _FILE_SHARE_WRITE,
            None,
            _OPEN_EXISTING,
            0,
            None,
        )
        if handle == -1:
            raise ctypes.WinError(ctypes.get_last_error())

        # Zet Win32-handle → C-bestandsdescriptor → Python-bestandsobject om
        fd = msvcrt.open_osfhandle(handle, os.O_RDONLY | os.O_BINARY)
        with open(fd, "rb") as fsrc, open(dst, "wb") as fdst:
            while chunk := fsrc.read(1 << 20):  # blokken van 1 MB
                fdst.write(chunk)
    except OSError:
        # Laatste redmiddel: normale kopie (kan mislukken als de vergrendeling exclusief is)
        shutil.copy2(src, dst)


def start_background_refresh(excel_path: str | Path, source_path: str | Path | None = None) -> None:
    """
    Start een refresh op de achtergrond. Veilig om vanuit de hoofdthread aan te roepen; keert meteen terug.
    source_path: indien opgegeven (bijv. OneDrive-kopie), wordt gebruikt als snelle kopiebron.
    """
    global _refresh_thread, _refresh_error
    _refresh_done.clear()
    _refresh_error = ""

    def _run():
        global _refresh_error
        try:
            refresh_excel(excel_path, source_path=source_path)
        except Exception as exc:
            _refresh_error = str(exc)
        finally:
            _refresh_done.set()

    _refresh_thread = threading.Thread(target=_run, daemon=True, name="sp-refresh")
    _refresh_thread.start()


def _get_excel_proc_handle(xl) -> object | None:
    """
    Geef een Win32-proceshandle terug voor de Excel-instantie, of None bij mislukking.
    Gebruikt xl.ProcessID (COM-eigenschap) — betrouwbaar, ook voor instanties zonder venster.
    Valt terug op de Hwnd-aanpak voor oudere Excel-versies.
    """
    try:
        import win32api
        import win32con
        # xl.ProcessID is beschikbaar op Excel 2007+ en werkt voor verborgen instanties
        try:
            pid = xl.ProcessID
        except Exception:
            import win32process
            hwnd = xl.Hwnd
            if not hwnd:
                return None
            _, pid = win32process.GetWindowThreadProcessId(hwnd)
        return win32api.OpenProcess(
            win32con.SYNCHRONIZE | win32con.PROCESS_TERMINATE, False, pid
        )
    except Exception:
        return None


def _kill_excel_process(proc_handle) -> None:
    """
    Wacht tot 15 s totdat het Excel-proces netjes is afgesloten na Quit().
    Als het nog steeds leeft, beëindig het dan geforceerd en sluit de handle.
    """
    if proc_handle is None:
        time.sleep(5)   # terugval: blind wachten (verborgen Excel, geen handle)
        return
    try:
        import win32api
        import win32event
        import win32con
        result = win32event.WaitForSingleObject(proc_handle, 15_000)
        if result != win32con.WAIT_OBJECT_0:   # niet op tijd afgesloten
            win32api.TerminateProcess(proc_handle, 0)
    except Exception:
        pass
    finally:
        try:
            import win32api as _w32
            _w32.CloseHandle(proc_handle)
        except Exception:
            pass


def refresh_excel(excel_path: str | Path, source_path: str | Path | None = None) -> None:
    """
    Werk het lokale SharePoint-overzichtsbestand bij.

    Strategie (in volgorde van prioriteit):
      1. Als *source_path* bestaat (bijv. een OneDrive-synckopie), kopieer het dan → *excel_path*
         en voer daarna een COM-refresh uit op de lokale kopie om Power Query-data bij te werken.
      2. Voer anders direct een COM-refresh uit op *excel_path*.

    Gooit een fout bij mislukking zodat de aanroeper die kan rapporteren.
    """
    dst = Path(excel_path).resolve()

    # ── Strategie 1: kopieer eerst een verse bron vanuit OneDrive ──────────
    if source_path:
        src = Path(source_path).resolve()
        if src.exists():
            _shared_copy(src, dst)

    # ── COM-refresh om Power Query uit te voeren ────────────────────────────
    if not dst.exists():
        raise FileNotFoundError(f"SharePoint file not found: {dst}")

    try:
        import pythoncom
        import win32com.client as win32
    except ImportError:
        return  # geen pywin32 — refresh stilzwijgend overslaan

    # Gebruik een uniek tijdelijk bestand om conflicten met oude locks van eerdere runs te voorkomen
    import uuid
    tmp = Path(tempfile.gettempdir()) / f"_sp_refresh_{uuid.uuid4().hex[:8]}_{dst.name}"
    _shared_copy(dst, tmp)
    tmp_path = str(tmp)

    pythoncom.CoInitialize()
    try:
        try:
            xl = win32.GetActiveObject("Excel.Application")
            we_started_excel = False
            proc_handle = None
        except Exception:
            xl = win32.DispatchEx("Excel.Application")
            we_started_excel = True
            proc_handle = _get_excel_proc_handle(xl)
            xl.Visible = False

        prev_screen  = xl.ScreenUpdating
        prev_alerts  = xl.DisplayAlerts
        prev_events  = xl.EnableEvents
        xl.ScreenUpdating = False
        xl.DisplayAlerts  = False
        xl.EnableEvents   = False
        xl.Interactive    = False
        try:
            wb = xl.Workbooks.Open(tmp_path, UpdateLinks=0, ReadOnly=False)
            wb.RefreshAll()
            xl.CalculateUntilAsyncQueriesDone()
            wb.Save()
            wb.Close(SaveChanges=False)
        finally:
            xl.ScreenUpdating = prev_screen
            xl.DisplayAlerts  = prev_alerts
            xl.EnableEvents   = prev_events
            xl.Interactive    = True
            if we_started_excel:
                xl.Quit()
                _kill_excel_process(proc_handle)

        # Kopieer ververste data terug — probeer opnieuw als OneDrive dst kort vergrendelt
        for attempt in range(3):
            try:
                shutil.copy2(tmp, dst)
                break
            except PermissionError:
                time.sleep(2)
        else:
            raise PermissionError(f"Cannot write refreshed data back to {dst}")
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass


def read_deals(excel_path: str | Path) -> list[dict]:
    """
    Lees alle rijen uit het SharePoint-overzichtstabblad.
    Geeft een lijst met ruwe dicts terug, gesleuteld op kolomkop.
    """
    src = Path(excel_path)
    if not src.exists():
        raise FileNotFoundError(f"SharePoint summary not found: {src}")

    # Kopieer naar een tijdelijk bestand om OneDrive-/Excel-bestandslocks te vermijden
    tmp = Path(tempfile.gettempdir()) / f"sp_summary_{src.stem}.xlsx"
    _shared_copy(src, tmp)

    try:
        wb = openpyxl.load_workbook(tmp, data_only=True, read_only=True)
        try:
            ws = wb[SHEET_NAME]
        except KeyError:
            raise ValueError(
                f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"
            )
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    finally:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass

    if not rows:
        return []

    # Verwijder omringende aanhalingstekens die Power Query soms aan kopteksten toevoegt
    headers = [
        str(h).strip('"').strip() if h is not None else "" for h in rows[0]
    ]

    return [
        {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        for row in rows[1:]
    ]
